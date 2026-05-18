"""
Excel audit workbook: raw CSV values vs recap calculations (Key Group Performance, per-store).

Requires: openpyxl
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional


def _si(v: Any) -> int:
    try:
        return int(float(str(v).strip()))
    except Exception:
        return 0


def write_audit_workbook(
    *,
    output_xlsx_path: str,
    group_name: str,
    period_label: str,
    ib_rows: List[Dict[str, str]],
    ob_rows: Optional[List[Dict[str, str]]],
    ib_only: bool,
    sn_ordered: List[str],
    dd: Dict[str, Dict[str, Any]],
    inbound_has_soft_appt_column: bool,
    ib_opportunities_column: str,
    ob_opportunities_column: Optional[str] = None,
    ob_connected_column: str = "Connected",
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    def soft_ib(row: Dict[str, str]) -> int:
        if inbound_has_soft_appt_column:
            raw_s = str(row.get("Soft Appt", "")).strip()
            if raw_s != "":
                return _si(raw_s)
        tot = _si(row.get("Total Appts"))
        h = _si(row.get("Hard Appt"))
        return max(0, tot - h)

    def pct_appt_1dp(total_appts: int, denom: int) -> Optional[float]:
        if not denom:
            return None
        return round(100.0 * total_appts / denom, 1)

    def pct_hard_of_hard_plus_soft(hard: int, soft: int) -> int:
        denom = hard + soft
        return round(100.0 * hard / denom) if denom else 0

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Overview"
    ws0["A1"] = "PromptPath Executive Report — number audit"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"] = group_name
    ws0["A3"] = period_label
    ws0["A5"] = "Appt set rate"
    ws0["A5"].font = Font(bold=True)
    ib_desc = (
        "IB appointment set rate (Key Group Performance + store summaries) = inbound Total Appts ÷ unique inbound "
        f'column "{ib_opportunities_column}", one decimal.'
    )
    if not ib_only:
        ib_desc += (
            f' Outbound connect count is read from "{ob_connected_column}" (matches Connect-rate numerator '
            'and denominator for outbound appointment set rate). '
        )
        if ob_opportunities_column:
            ib_desc += f'Outbound Opportunities use unique column "{ob_opportunities_column}".'
    ws0["A6"] = ib_desc
    ws0["A7"] = (
        '% hard = Hard Appt ÷ (Hard Appt + Soft Appt), rounded to whole percent. '
        'Soft uses CSV "Soft Appt" when present; otherwise Soft = Total Appts − Hard Appt '
        '(appointment counts only — not inbound call volume).'
    )
    ws0["A8"] = (
        '"Sum of dealer rows" sums each dealership row from the inbound CSV (excluding All Dealers). '
        'Compare to "All Dealers" to see rollup vs arithmetic sum differences.'
    )
    ws0["A9"] = f"Inbound Opportunities column used (unique customer): {ib_opportunities_column}"
    if ob_opportunities_column:
        ws0["A10"] = f"Outbound Opportunities column used (unique customer): {ob_opportunities_column}"
    ws0.column_dimensions["A"].width = 100

    ws1 = wb.create_sheet("Inbound_KGP_vs_sum")
    h1 = [
        "Row label",
        "Period",
        "Inbound Calls",
        "Connected",
        f"Unique ({ib_opportunities_column})",
        "Total Appts",
        "Hard Appt",
        "Soft (CSV or Total−Hard)",
        "IB appt set rate % (÷ unique)",
        "% hard (Hard÷Hard+Soft)",
        "Check: Hard+Soft",
    ]
    for c, x in enumerate(h1, 1):
        ws1.cell(1, c, value=x).font = Font(bold=True)

    def dealer_ib_rows(period: str) -> List[Dict[str, str]]:
        out = []
        for r in ib_rows:
            if r.get("Dealerships", "").strip() == "All Dealers":
                continue
            if r.get("Period", "").strip() != period:
                continue
            out.append(r)
        return out

    def sum_ag(rows: List[Dict[str, str]]) -> Dict[str, int]:
        ibc = sum(_si(r.get("Inbound Calls")) for r in rows)
        con = sum(_si(r.get("Connected")) for r in rows)
        tap = sum(_si(r.get("Total Appts")) for r in rows)
        har = sum(_si(r.get("Hard Appt")) for r in rows)
        sof = sum(soft_ib(r) for r in rows)
        uni = sum(_si(r.get(ib_opportunities_column)) for r in rows)
        return {"ib_calls": ibc, "connected": con, "uniq_opps": uni, "total_appts": tap, "hard": har, "soft": sof}

    def write_row(ridx: int, label: str, period: str, row: Optional[Dict[str, str]]) -> None:
        if not row:
            ws1.cell(ridx, 1, value=f"{label} ({period})")
            ws1.cell(ridx, 2, value=period)
            for j in range(3, 12):
                ws1.cell(ridx, j, value="—")
            return
        ic = _si(row.get("Connected"))
        iu = _si(row.get(ib_opportunities_column))
        ia = _si(row.get("Total Appts"))
        ih = _si(row.get("Hard Appt"))
        isoft = soft_ib(row)
        chk = ih + isoft
        rate = pct_appt_1dp(ia, iu)
        hp = pct_hard_of_hard_plus_soft(ih, isoft)
        ws1.cell(ridx, 1, value=label)
        ws1.cell(ridx, 2, value=period)
        ws1.cell(ridx, 3, value=_si(row.get("Inbound Calls")))
        ws1.cell(ridx, 4, value=ic)
        ws1.cell(ridx, 5, value=iu)
        ws1.cell(ridx, 6, value=ia)
        ws1.cell(ridx, 7, value=ih)
        ws1.cell(ridx, 8, value=isoft)
        ws1.cell(ridx, 9, value=rate)
        ws1.cell(ridx, 10, value=hp)
        ws1.cell(ridx, 11, value=chk)

    r = 2
    ad_curr = next(
        (x for x in ib_rows if x.get("Dealerships", "").strip() == "All Dealers" and x.get("Period", "").strip() == "Current"),
        None,
    )
    ad_prev = next(
        (x for x in ib_rows if x.get("Dealerships", "").strip() == "All Dealers" and x.get("Period", "").strip() == "Previous"),
        None,
    )
    write_row(r, "All Dealers (recap KGP uses this)", "Current", ad_curr)
    r += 1
    write_row(r, "All Dealers", "Previous", ad_prev)
    r += 1

    for period in ("Current", "Previous"):
        rows = dealer_ib_rows(period)
        if not rows:
            continue
        agg = sum_ag(rows)
        ih, isoft = agg["hard"], agg["soft"]
        chk = ih + isoft
        rate = pct_appt_1dp(agg["total_appts"], agg["uniq_opps"])
        hp = pct_hard_of_hard_plus_soft(ih, isoft)
        ws1.cell(r, 1, value=f"Sum of dealer rows ({len(rows)} dealers)")
        ws1.cell(r, 2, value=period)
        ws1.cell(r, 3, value=agg["ib_calls"])
        ws1.cell(r, 4, value=agg["connected"])
        ws1.cell(r, 5, value=agg["uniq_opps"])
        ws1.cell(r, 6, value=agg["total_appts"])
        ws1.cell(r, 7, value=ih)
        ws1.cell(r, 8, value=isoft)
        ws1.cell(r, 9, value=rate)
        ws1.cell(r, 10, value=hp)
        ws1.cell(r, 11, value=chk)
        r += 1

    ws1.column_dimensions["A"].width = 40
    for col in "BCDEFGHIJK":
        ws1.column_dimensions[col].width = 14

    ws2 = wb.create_sheet("Stores_in_report")
    h2 = [
        "Dealership",
        "curr_ib_unique_opps",
        "curr_connected_IB",
        "curr_ob_unique_opps",
        "curr_connected_OB",
        "curr_total_appts",
        "curr_hard",
        "curr_soft",
        "IB appt set rate % (÷ unique)",
        "% hard",
    ]
    if not ib_only:
        h2.append("OB appt set rate % (÷ connected)")
    for c, x in enumerate(h2, 1):
        ws2.cell(1, c, value=x).font = Font(bold=True)
    rr = 2
    for name in sn_ordered:
        d = dd.get(name, {})
        conn = int(d.get("curr_connected", 0))
        tap = int(d.get("curr_total_appts", 0))
        hap = int(d.get("curr_hard_appts", 0))
        sap = d.get("curr_soft_appts")
        if sap is None:
            sap = max(0, tap - hap)
        else:
            sap = int(sap)
        iuo = int(d.get("curr_ib_unique_opps", 0))
        ouo = int(d.get("curr_ob_unique_opps", 0)) if not ib_only else ""
        obconn = int(d.get("curr_ob_connected", 0)) if not ib_only else ""
        ws2.cell(rr, 1, value=name)
        ws2.cell(rr, 2, value=iuo)
        ws2.cell(rr, 3, value=conn)
        ws2.cell(rr, 4, value=ouo if not ib_only else "")
        ws2.cell(rr, 5, value=obconn if not ib_only else "")
        ws2.cell(rr, 6, value=tap)
        ws2.cell(rr, 7, value=hap)
        ws2.cell(rr, 8, value=sap)
        ws2.cell(rr, 9, value=pct_appt_1dp(tap, iuo))
        ws2.cell(rr, 10, value=pct_hard_of_hard_plus_soft(hap, sap))
        if not ib_only:
            tot_ob = int(d.get("curr_ob_total_appts", 0))
            ws2.cell(rr, 11, value=pct_appt_1dp(tot_ob, int(obconn) if obconn != "" else 0))
        rr += 1

    if not ib_only and ob_rows:
        ws3 = wb.create_sheet("Outbound_All_Dealers")
        ho = [
            "Dealerships",
            "Period",
            "Outbound Dials",
            ob_connected_column,
            f"Unique ({ob_opportunities_column})",
            "Hard Appt",
            "Soft Appt",
            "Hard+Soft",
            "OB appt set rate % (÷ connected)",
        ]
        for c, x in enumerate(ho, 1):
            ws3.cell(1, c, value=x).font = Font(bold=True)
        r3 = 2
        for period in ("Current", "Previous"):
            row = next(
                (
                    x
                    for x in ob_rows
                    if x.get("Dealerships", "").strip() == "All Dealers" and x.get("Period", "").strip() == period
                ),
                None,
            )
            if not row:
                continue
            conn = _si(row.get(ob_connected_column))
            ou = _si(row.get(ob_opportunities_column)) if ob_opportunities_column else conn
            ha = _si(row.get("Hard Appt"))
            sa = _si(row.get("Soft Appt"))
            tot = ha + sa
            ws3.cell(r3, 1, value="All Dealers")
            ws3.cell(r3, 2, value=period)
            ws3.cell(r3, 3, value=_si(row.get("Outbound Dials")))
            ws3.cell(r3, 4, value=conn)
            ws3.cell(r3, 5, value=ou)
            ws3.cell(r3, 6, value=ha)
            ws3.cell(r3, 7, value=sa)
            ws3.cell(r3, 8, value=tot)
            ws3.cell(r3, 9, value=pct_appt_1dp(tot, conn))
            r3 += 1

    Path(output_xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path
